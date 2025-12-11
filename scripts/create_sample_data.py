#!/usr/bin/env python3
"""
Create sample demo data for the Team Feedback Tool.
Generates fictitious orgchart and feedback data for demos and testing.

Usage:
    python3 create_sample_data.py                 # Creates small team orgchart CSV only
    python3 create_sample_data.py --large         # Creates large org orgchart CSV only
    python3 create_sample_data.py --demo          # Full demo setup (recommended)
    python3 create_sample_data.py --large --demo  # Large org full demo

The --demo flag:
    1. Generates orgchart CSV
    2. Imports orgchart to database
    3. Generates peer feedback in database
    4. Generates manager feedback in database
    5. Generates Workday XLSX with mix of structured/generic feedback

Output:
    - samples/sample-orgchart.csv (or sample-orgchart-large.csv)
    - With --demo:
        - feedback.db populated with sample data
        - samples/sample-workday-feedback.xlsx (Workday format with [TENETS] markers)
"""

import csv
import os
import random
import sys
import json
from datetime import datetime, timedelta

SAMPLES_DIR = 'samples'
from collections import defaultdict
from feedback_models import init_db, Feedback, ManagerFeedback

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_user_id(name):
    """Generate user ID from name (first initial + last name, lowercase)"""
    parts = name.split()
    if len(parts) >= 2:
        return (parts[0][0] + parts[-1]).lower().replace("'", "").replace(".", "")
    return name.lower().replace(" ", "")[:8]


def generate_email(user_id):
    """Generate email from user ID"""
    return f"{user_id}@example.com"


def get_location():
    """Get random location"""
    locations = [
        'Remote US CA', 'Remote US TX', 'Remote US NY', 'Remote US MA',
        'Raleigh NC', 'Boston MA', 'San Francisco CA',
        'Remote UK', 'Remote Ireland', 'Remote France', 'Remote Spain',
        'Brno CZ', 'Pune IN'
    ]
    return random.choice(locations)


def get_small_team_data():
    """
    Small team: 12 employees under single manager (Della Gate).
    Perfect for testing with a manageable dataset.
    """
    manager = 'dgate'

    employees = [
        ('Paige Duty', 'Staff SRE'),
        ('Lee Latency', 'Senior Software Developer'),
        ('Mona Torr', 'Senior SRE'),
        ('Robin Rollback', 'Software Developer'),
        ('Kenny Canary', 'Software Developer'),
        ('Tracey Loggins', 'Senior SRE'),
        ('Sue Q. Ell', 'Senior Software Developer'),
        ('Jason Blob', 'Software Developer'),
        ('Al Ert', 'Staff SRE'),
        ('Addie Min', 'Senior Software Developer'),
        ('Tim Out', 'Software Developer'),
        ('Barbie Que', 'Senior SRE'),
    ]

    result = []
    for name, job in employees:
        user_id = generate_user_id(name)
        result.append({
            'name': name,
            'user_id': user_id,
            'job_title': job,
            'location': get_location(),
            'email': generate_email(user_id),
            'manager_uid': manager
        })

    # Add the manager
    result.append({
        'name': 'Della Gate',
        'user_id': 'dgate',
        'job_title': 'Engineering Manager',
        'location': 'Raleigh NC',
        'email': 'dgate@example.com',
        'manager_uid': ''  # Top-level
    })

    return result


def get_large_org_data():
    """
    Large org: 50 employees across 5 managers.
    Tests multi-manager scenario.
    """
    managers = {
        'dgate': 'Della Gate',
        'rmap': 'Rhoda Map',
        'keye': 'Kay P. Eye',
        'aenda': 'Agie Enda',
        'mstone': 'Mai Stone'
    }

    names = [
        'Paige Duty', 'Lee Latency', 'Mona Torr', 'Robin Rollback',
        'Kenny Canary', 'Tracey Loggins', 'Sue Q. Ell', 'Jason Blob',
        'Al Ert', 'Addie Min', 'Tim Out', 'Barbie Que',
        'Terry Byte', 'Nole Pointer', 'Marge Conflict', 'Bridget Branch',
        'Cody Ryder', 'Cy Ferr', 'Phil Wall', 'Lana Wan',
        'Artie Ficial', 'Ruth Cause', 'Matt Rick', 'Cassie Cache',
        'Sue Do', 'Pat Ch', 'Devin Null', 'Justin Time',
        'Annie O\'Maly', 'Sam Box', 'Val Idation', 'Bill Ding',
        'Ty Po', 'Mike Roservices', 'Lou Pe', 'Connie Tainer',
        'Noah Node', 'Sara Ver', 'Exa M. Elle', 'Dee Ploi',
        'Ray D. O\'Button', 'Cam Elcase', 'Hashim Map', 'Ben Chmark',
        'Grace Full', 'Shel Script', 'Sal T. Hash', 'Reba Boot',
        'Stan Dup', 'Kay Eight'
    ]

    jobs = [
        'Principal Software Developer', 'Staff Software Developer',
        'Senior Software Developer', 'Software Developer',
        'Principal SRE', 'Staff SRE', 'Senior SRE', 'SRE',
        'Engineering Manager'
    ]

    result = []
    name_idx = 0

    # Create employees for each manager (10 per manager)
    for manager_uid in managers.keys():
        for i in range(10):
            if name_idx >= len(names):
                break
            name = names[name_idx]
            user_id = generate_user_id(name)
            job = random.choice(jobs)

            result.append({
                'name': name,
                'user_id': user_id,
                'job_title': job,
                'location': get_location(),
                'email': generate_email(user_id),
                'manager_uid': manager_uid
            })
            name_idx += 1

    # Add managers
    for manager_uid, manager_name in managers.items():
        result.append({
            'name': manager_name,
            'user_id': manager_uid,
            'job_title': 'Engineering Manager',
            'location': get_location(),
            'email': generate_email(manager_uid),
            'manager_uid': ''  # Top-level
        })

    return result


def write_orgchart_csv(filename, people):
    """Write orgchart CSV file"""
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Name', 'User ID', 'Job Title', 'Location', 'Email', 'Manager UID'])

        for person in people:
            writer.writerow([
                person['name'],
                person['user_id'],
                person['job_title'],
                person['location'],
                person['email'],
                person['manager_uid']
            ])

    print(f"✓ Created {filename} with {len(people)} people")


def generate_sample_feedback(people):
    """
    Generate realistic sample feedback in database with ~80% coverage.
    Returns list of feedback dicts for CSV export.
    """
    # Load tenets
    with open('samples/tenets-sample.json', 'r') as f:
        tenets_data = json.load(f)
    tenets = [t for t in tenets_data['tenets'] if t.get('active', True)]
    tenet_ids = [t['id'] for t in tenets]

    session = init_db()

    # Clear existing feedback
    session.query(Feedback).delete()

    # Get employees (not managers)
    employees = [p for p in people if p['manager_uid']]
    people_by_id = {p['user_id']: p for p in people}

    feedback_list = []  # For CSV export

    # Varied feedback text templates
    strength_texts = [
        "{name} consistently demonstrates excellence in these areas and serves as a role model for the team.",
        "I've observed {name} excel in these tenets throughout our collaboration this quarter.",
        "{name} brings exceptional capability in these areas, making significant positive impact.",
        "These are standout strengths for {name}. They handle challenges in these areas with expertise and professionalism.",
        "Working with {name} has shown me their strong command of these principles. Keep up the great work!",
        "{name} sets the bar high in these tenets and helps elevate the entire team's performance.",
        "I appreciate how {name} embodies these values in their daily work. It makes a real difference.",
        "{name} has shown consistent growth and mastery in these areas over the time we've worked together.",
        "These tenets are clearly where {name} shines brightest. Their expertise is invaluable to our success.",
        "I've been impressed by {name}'s dedication to these principles and the results they achieve.",
        "{name} demonstrates deep understanding and application of these tenets in everything they do.",
        "The team benefits greatly from {name}'s strength in these areas. Excellent work overall.",
    ]

    improvement_texts = [
        "I see opportunities for {name} to develop further in these areas to increase their overall impact.",
        "These tenets could use more attention and focus from {name} going forward.",
        "With some dedicated effort in these areas, {name} could take their contributions to the next level.",
        "I'd encourage {name} to prioritize growth in these tenets as they continue their development.",
        "These are areas where I think {name} has room to grow and strengthen their skillset.",
        "Focusing on these principles would help {name} become even more effective in their role.",
        "I believe {name} would benefit from additional practice and experience with these tenets.",
        "These areas represent growth opportunities that could enhance {name}'s overall effectiveness.",
        "I'd like to see {name} invest more time and energy into developing these capabilities.",
        "While {name} has many strengths, these tenets could use some improvement and refinement.",
        "Working on these areas would round out {name}'s already solid skillset nicely.",
        "I think {name} could make meaningful progress by focusing attention on these tenets.",
    ]

    # Generate feedback: each employee gives feedback to ~80% of peers
    for employee in employees:
        # Find peers (same manager)
        peers = [p for p in employees
                 if p['user_id'] != employee['user_id']
                 and p['manager_uid'] == employee['manager_uid']]

        # Calculate 80% of peers (minimum 1, maximum all peers)
        num_feedback = max(1, int(len(peers) * 0.8))
        num_feedback = min(num_feedback, len(peers))

        # Randomly select which peers to give feedback to
        selected_peers = random.sample(peers, num_feedback) if len(peers) > 0 else []

        for peer in selected_peers:
            # Select 3 random strengths
            strengths = random.sample(tenet_ids, 3)

            # Select 2-3 random improvements
            num_improvements = random.choice([2, 3])
            available_improvements = [tid for tid in tenet_ids if tid not in strengths]
            improvements = random.sample(available_improvements, num_improvements)

            # Generate sample text with name substitution
            strength_text = random.choice(strength_texts).format(name=peer['name'])
            improvement_text = random.choice(improvement_texts).format(name=peer['name'])

            feedback = Feedback(
                from_user_id=employee['user_id'],
                to_user_id=peer['user_id'],
                strengths_text=strength_text,
                improvements_text=improvement_text
            )
            feedback.set_strengths(strengths)
            feedback.set_improvements(improvements)

            session.add(feedback)

            # Also save for CSV export
            feedback_list.append({
                'from_user_id': employee['user_id'],
                'to_user_id': peer['user_id'],
                'to_manager_uid': peer['manager_uid'],
                'strengths': strengths,
                'improvements': improvements,
                'strengths_text': strength_text,
                'improvements_text': improvement_text
            })

    session.commit()
    session.close()

    # Calculate coverage statistics
    team_sizes = defaultdict(int)
    for emp in employees:
        team_sizes[emp['manager_uid']] += 1

    total_possible = sum(size * (size - 1) for size in team_sizes.values())
    coverage_percent = (len(feedback_list) / total_possible * 100) if total_possible > 0 else 0

    print(f"✓ Generated {len(feedback_list)} peer feedback entries")
    print(f"  Coverage: {coverage_percent:.1f}% of possible peer feedback within teams")

    return feedback_list


def generate_manager_feedback(people):
    """
    Generate manager feedback records with highlighted tenets and commentary.
    Analyzes peer feedback to select top tenets for each team member.
    """
    session = init_db()

    # Clear existing manager feedback
    session.query(ManagerFeedback).delete()

    # Get managers and their teams
    managers = [p for p in people if not p['manager_uid']]
    employees = [p for p in people if p['manager_uid']]

    # Manager feedback text templates
    manager_texts = [
        "Based on peer feedback, {name} shows strong performance in the highlighted areas. "
        "Continue developing the improvement areas identified.",
        "{name} has received consistent positive feedback from peers. "
        "The highlighted tenets reflect team observations.",
        "Peer feedback confirms {name}'s strengths align with team expectations. "
        "Focus on the improvement areas for continued growth.",
        "{name} demonstrates solid capabilities as reflected in peer feedback. "
        "The highlighted areas represent key themes from multiple reviewers.",
        "Team feedback highlights {name}'s contributions and growth opportunities. "
        "Continue leveraging strengths while addressing improvement areas.",
    ]

    count = 0
    for manager in managers:
        # Get team members
        team = [e for e in employees if e['manager_uid'] == manager['user_id']]

        for member in team:
            # Get peer feedback for this member
            feedbacks = session.query(Feedback).filter_by(to_user_id=member['user_id']).all()

            if not feedbacks:
                continue

            # Aggregate tenet counts
            strength_counts = defaultdict(int)
            improvement_counts = defaultdict(int)

            for fb in feedbacks:
                for s in fb.get_strengths():
                    strength_counts[s] += 1
                for i in fb.get_improvements():
                    improvement_counts[i] += 1

            # Select top 2-3 strengths and 1-2 improvements
            top_strengths = sorted(strength_counts.keys(),
                                   key=lambda x: strength_counts[x],
                                   reverse=True)[:random.choice([2, 3])]
            top_improvements = sorted(improvement_counts.keys(),
                                      key=lambda x: improvement_counts[x],
                                      reverse=True)[:random.choice([1, 2])]

            # Generate manager commentary
            feedback_text = random.choice(manager_texts).format(name=member['name'])

            mgr_feedback = ManagerFeedback(
                manager_uid=manager['user_id'],
                team_member_uid=member['user_id'],
                feedback_text=feedback_text
            )
            mgr_feedback.set_selected_strengths(top_strengths)
            mgr_feedback.set_selected_improvements(top_improvements)

            session.add(mgr_feedback)
            count += 1

    session.commit()
    session.close()

    print(f"✓ Generated {count} manager feedback entries")


def generate_workday_xlsx(feedback_list, people, include_structured=True):
    """
    Generate a sample Workday "Feedback on My Team" XLSX export.
    Simulates both structured (tool-assisted) and generic feedback.

    Args:
        feedback_list: List of feedback dicts from generate_sample_feedback
        people: List of all people dicts
        include_structured: If True, include [TENETS] markers in some feedback
    """
    if not HAS_OPENPYXL:
        print("⚠ openpyxl not installed, skipping XLSX generation")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback on My Team"

    # Write header rows
    ws['A1'] = 'Feedback Received'
    ws['A2'] = 'About Photo'
    ws['B2'] = 'About'
    ws['C2'] = 'Feedback Also Given To'
    ws['D2'] = 'From Photo'
    ws['E2'] = 'From'
    ws['F2'] = 'Question'
    ws['G2'] = 'Feedback'
    ws['H2'] = 'Asked By'
    ws['I2'] = 'Type'
    ws['J2'] = 'Date'

    # Get people lookup
    people_by_id = {p['user_id']: p for p in people}
    managers = {p['user_id']: p for p in people if not p['manager_uid']}

    # Load tenets for structured feedback format
    with open('samples/tenets-sample.json', 'r') as f:
        tenets_data = json.load(f)
    tenets = {t['id']: t['name'] for t in tenets_data['tenets'] if t.get('active', True)}

    # Workday feedback question templates
    questions = [
        "What strengths does the associate demonstrate?",
        "What should the associate focus on to continue to develop?",
        "Please provide feedback on the associate's performance."
    ]

    # Generic feedback text templates (for non-structured entries)
    generic_feedback_templates = [
        "{name} has been a great team player this quarter. Strong communication and reliable delivery.",
        "Good collaboration skills. {name} is always willing to help others.",
        "{name} demonstrates solid technical skills and ownership of their work.",
        "I've enjoyed working with {name}. They bring positive energy to the team.",
        "{name} is dependable and delivers quality work consistently.",
        "{name} could improve by being more proactive in sharing updates.",
        "Solid contributor. {name} handles challenges well and stays focused.",
        "{name} is responsive and easy to work with on projects."
    ]

    row = 3  # Start data after headers
    base_date = datetime.now() - timedelta(days=30)

    # Generate feedback entries
    for fb in feedback_list:
        to_person = people_by_id.get(fb['to_user_id'])
        from_person = people_by_id.get(fb['from_user_id'])

        if not to_person or not from_person:
            continue

        manager = managers.get(to_person['manager_uid'])
        if not manager:
            continue

        # Random date within last 60 days
        feedback_date = base_date + timedelta(days=random.randint(-30, 30))

        # Decide if this is self-requested or manager-requested
        is_self_request = random.choice([True, True, False])  # 2/3 self-requested
        asked_by = to_person['name'] if is_self_request else manager['name']
        request_type = "Requested by Self" if is_self_request else "Requested by Others"

        # Decide if structured (with tenets) or generic
        if include_structured and random.random() < 0.6:  # 60% structured
            # Generate structured feedback with [TENETS] marker at bottom
            strength_names = [tenets.get(s, s) for s in fb['strengths']]
            improvement_names = [tenets.get(i, i) for i in fb['improvements']]

            feedback_text = f"""Strengths:
{chr(10).join('• ' + name for name in strength_names)}

{fb['strengths_text']}

Areas for Improvement:
{chr(10).join('• ' + name for name in improvement_names)}

{fb['improvements_text']}

[TENETS]
Strengths: {', '.join(fb['strengths'])}
Improvements: {', '.join(fb['improvements'])}
[/TENETS]"""
        else:
            # Generate generic feedback
            feedback_text = random.choice(generic_feedback_templates).format(name=to_person['name'])

        # Pick a question
        question = random.choice(questions)

        # Write row
        ws[f'A{row}'] = ''  # About Photo (empty)
        ws[f'B{row}'] = to_person['name']
        ws[f'C{row}'] = ''  # Feedback Also Given To (empty)
        ws[f'D{row}'] = ''  # From Photo (empty)
        ws[f'E{row}'] = from_person['name']
        ws[f'F{row}'] = question
        ws[f'G{row}'] = feedback_text
        ws[f'H{row}'] = asked_by
        ws[f'I{row}'] = request_type
        ws[f'J{row}'] = feedback_date

        row += 1

    # Also add some generic entries that aren't from our feedback_list
    # to simulate other Workday feedback workflows
    for _ in range(min(10, len(people) // 3)):
        to_person = random.choice([p for p in people if p['manager_uid']])
        from_person = random.choice([p for p in people if p['user_id'] != to_person['user_id']])
        manager = managers.get(to_person['manager_uid'])

        if not manager:
            continue

        feedback_date = base_date + timedelta(days=random.randint(-60, 0))
        asked_by = to_person['name']
        request_type = "Requested by Self"

        feedback_text = random.choice(generic_feedback_templates).format(name=to_person['name'])
        question = "Please provide feedback on the associate's performance."

        ws[f'A{row}'] = ''
        ws[f'B{row}'] = to_person['name']
        ws[f'C{row}'] = ''
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = from_person['name']
        ws[f'F{row}'] = question
        ws[f'G{row}'] = feedback_text
        ws[f'H{row}'] = asked_by
        ws[f'I{row}'] = request_type
        ws[f'J{row}'] = feedback_date

        row += 1

    # Save file
    filename = os.path.join(SAMPLES_DIR, 'sample-workday-feedback.xlsx')
    wb.save(filename)

    total_rows = row - 3
    print(f"✓ Generated {filename} with {total_rows} feedback entries")

    return filename


def main():
    large = '--large' in sys.argv
    demo = '--demo' in sys.argv

    # Ensure samples directory exists
    os.makedirs(SAMPLES_DIR, exist_ok=True)

    # Generate people data
    if large:
        people = get_large_org_data()
        filename = os.path.join(SAMPLES_DIR, 'sample-orgchart-large.csv')
    else:
        people = get_small_team_data()
        filename = os.path.join(SAMPLES_DIR, 'sample-orgchart.csv')

    # Write orgchart CSV
    write_orgchart_csv(filename, people)

    if demo:
        # Import orgchart to database
        print("\nImporting orgchart to database...")
        from import_orgchart import import_orgchart
        import_orgchart(filename)

        # Generate peer feedback
        print("\nGenerating sample feedback...")
        feedback_list = generate_sample_feedback(people)

        # Generate manager feedback
        print("\nGenerating manager feedback...")
        generate_manager_feedback(people)

        # Generate Workday XLSX
        print("\nGenerating Workday XLSX...")
        xlsx_file = generate_workday_xlsx(feedback_list, people)

        print("\n" + "=" * 50)
        print("Demo setup complete!")
        print("=" * 50)
        print("\nRun the app:  python3 app.py")
        print("Then visit:   http://localhost:5001")
        if xlsx_file:
            print(f"\nTo test Workday import:")
            print(f"  1. Go to Manager mode and select a manager")
            print(f"  2. Import {xlsx_file}")
    else:
        print("\nNext steps:")
        print(f"  1. Import orgchart: python3 import_orgchart.py {filename}")
        print(f"  2. Start app: python3 app.py")
        print(f"\nOr use --demo for a complete setup with sample feedback")


if __name__ == '__main__':
    if '--help' in sys.argv or '-h' in sys.argv:
        print(__doc__)
        sys.exit(0)

    main()
