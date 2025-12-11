"""
Import feedback from Workday XLSX export.

Parses the "Feedback on My Team" export from Workday and imports
feedback entries into the database. Distinguishes between:
1. Structured feedback (tool-assisted with [TENETS] marker)
2. Generic feedback (free-text from other WD workflows)

Column mappings are configurable via workday_config.json to support
different Workday export formats.
"""

import openpyxl
import json
import os
from datetime import datetime
from sqlalchemy.exc import IntegrityError
from feedback_models import init_db, WorkdayFeedback


# Default column header names (case-insensitive matching)
# Maps internal field names to expected column header text
DEFAULT_CONFIG = {
    "column_headers": {
        "about": ["about", "recipient", "employee", "for"],
        "from_name": ["from", "provider", "given by", "reviewer"],
        "question": ["question"],
        "feedback": ["feedback", "response", "answer", "comments"],
        "asked_by": ["asked by", "requested by", "requester"],
        "request_type": ["type", "request type"],
        "date": ["date", "response date", "submitted"]
    },
    "optional_headers": {
        "about_id": ["about id", "recipient id", "employee id"],
        "from_id": ["from id", "provider id", "reviewer id"],
        "feedback_also_given_to": ["feedback also given to", "also given to"]
    },
    "header_row": 2,
    "request_types": {
        "self": "Requested by Self",
        "others": "Requested by Others"
    }
}


def load_config():
    """Load Workday import configuration from workday_config.json if it exists."""
    config_path = os.path.join(os.path.dirname(__file__), 'workday_config.json')
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            return json.load(f)
    return DEFAULT_CONFIG


def detect_columns(ws, config):
    """Detect column positions from header row.

    Args:
        ws: openpyxl worksheet
        config: Import configuration dict

    Returns:
        Tuple of (col_mapping dict, warnings list)
        col_mapping maps field names to column indices (0-based)
    """
    header_row = config.get('header_row', 2)
    col_headers = config.get('column_headers', DEFAULT_CONFIG['column_headers'])
    opt_headers = config.get('optional_headers', DEFAULT_CONFIG.get('optional_headers', {}))

    # Read header row
    headers = []
    for cell in ws[header_row]:
        val = str(cell.value).lower().strip() if cell.value else ''
        headers.append(val)

    col_mapping = {}
    warnings = []

    # Match required columns - prefer exact matches over partial matches
    for field, possible_names in col_headers.items():
        found = False
        # First pass: look for exact matches
        for idx, header in enumerate(headers):
            if any(name.lower() == header for name in possible_names):
                col_mapping[field] = idx
                found = True
                break
        # Second pass: look for headers that start with the name (but not ones containing "photo")
        if not found:
            for idx, header in enumerate(headers):
                if 'photo' in header:
                    continue  # Skip photo columns
                if any(header.startswith(name.lower()) for name in possible_names):
                    col_mapping[field] = idx
                    found = True
                    break
        if not found:
            # Required field not found - this is a warning, we'll handle missing data gracefully
            col_mapping[field] = None

    # Match optional columns
    for field, possible_names in opt_headers.items():
        for idx, header in enumerate(headers):
            if any(name.lower() in header for name in possible_names):
                col_mapping[field] = idx
                break
        else:
            col_mapping[field] = None

    # Check for critical missing columns
    if col_mapping.get('about') is None:
        warnings.append("Could not find 'About'/'Recipient' column - using first non-photo column")
        # Try to find first text column that's not a photo
        for idx, header in enumerate(headers):
            if header and 'photo' not in header:
                col_mapping['about'] = idx
                break

    if col_mapping.get('from_name') is None:
        warnings.append("Could not find 'From'/'Provider' column - required for import")

    if col_mapping.get('feedback') is None:
        warnings.append("Could not find 'Feedback'/'Response' column")

    return col_mapping, warnings


def get_cell_value(row, col_idx):
    """Safely get a cell value from a row, returning None if index is out of range or None."""
    if col_idx is None:
        return None
    if col_idx >= len(row):
        return None
    return row[col_idx]


class ImportResult:
    """Result of an XLSX import operation."""

    def __init__(self):
        self.imported = 0
        self.skipped_duplicates = 0
        self.skipped_empty = 0
        self.structured_count = 0
        self.generic_count = 0
        self.warnings = []
        self.errors = []

    @property
    def success(self):
        return len(self.errors) == 0

    def to_dict(self):
        return {
            'success': self.success,
            'imported': self.imported,
            'skipped_duplicates': self.skipped_duplicates,
            'skipped_empty': self.skipped_empty,
            'structured_count': self.structured_count,
            'generic_count': self.generic_count,
            'warnings': self.warnings,
            'errors': self.errors
        }


def validate_row(row, row_num, col_mapping, config):
    """Validate a single row from the XLSX.

    Args:
        row: Tuple of cell values
        row_num: Row number for error messages
        col_mapping: Dict mapping field names to column indices
        config: Import configuration dict

    Returns:
        Tuple of (is_valid, error_message or None)
    """
    req_types = config.get('request_types', DEFAULT_CONFIG['request_types'])

    about = get_cell_value(row, col_mapping.get('about'))
    from_name = get_cell_value(row, col_mapping.get('from_name'))
    asked_by = get_cell_value(row, col_mapping.get('asked_by'))
    request_type = get_cell_value(row, col_mapping.get('request_type'))

    # Skip empty rows (section headers or pending requests)
    if not from_name:
        return False, None  # Not an error, just skip

    # Validate type consistency (only if all fields are present)
    if about and asked_by and request_type:
        if about == asked_by and request_type != req_types['self']:
            return False, (f"Row {row_num}: Data inconsistency - About '{about}' matches "
                          f"Asked By but Type is '{request_type}' (expected '{req_types['self']}')")
        if about != asked_by and request_type != req_types['others']:
            return False, (f"Row {row_num}: Data inconsistency - About '{about}' differs from "
                          f"Asked By '{asked_by}' but Type is '{request_type}' "
                          f"(expected '{req_types['others']}')")

    return True, None


def import_workday_xlsx(file_path, db_path='feedback.db', config=None):
    """Import feedback from a Workday XLSX export.

    Args:
        file_path: Path to the XLSX file
        db_path: Path to the SQLite database
        config: Optional import configuration (loads from workday_config.json if not provided)

    Returns:
        ImportResult with import statistics and any warnings/errors
    """
    result = ImportResult()

    # Load configuration
    if config is None:
        config = load_config()

    header_row = config.get('header_row', 2)

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        result.errors.append(f"Failed to open XLSX file: {e}")
        return result

    # Find the feedback sheet
    sheet_name = None
    for name in wb.sheetnames:
        if 'feedback' in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        # Use first sheet if no feedback sheet found
        sheet_name = wb.sheetnames[0]
        result.warnings.append(f"No 'Feedback' sheet found, using '{sheet_name}'")

    ws = wb[sheet_name]

    # Detect columns from header row
    col_mapping, col_warnings = detect_columns(ws, config)
    result.warnings.extend(col_warnings)

    # Check for critical missing columns
    if col_mapping.get('from_name') is None:
        result.errors.append("Cannot import: 'From'/'Provider' column not found in spreadsheet")
        return result

    if col_mapping.get('about') is None:
        result.errors.append("Cannot import: 'About'/'Recipient' column not found in spreadsheet")
        return result

    session = init_db(db_path)

    # Track if we see any data in "Feedback Also Given To" column
    feedback_also_given_to_used = False

    # Process rows (skip header rows)
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # Check for "Feedback Also Given To" usage
        feedback_also = get_cell_value(row, col_mapping.get('feedback_also_given_to'))
        if feedback_also:
            feedback_also_given_to_used = True

        # Validate row
        is_valid, error = validate_row(row, row_num, col_mapping, config)

        if error:
            result.errors.append(error)
            continue

        if not is_valid:
            # Empty row, skip silently but count
            result.skipped_empty += 1
            continue

        # Extract values using detected column mapping
        about = get_cell_value(row, col_mapping.get('about'))
        from_name = get_cell_value(row, col_mapping.get('from_name'))
        question = get_cell_value(row, col_mapping.get('question'))
        feedback_text = get_cell_value(row, col_mapping.get('feedback'))
        asked_by = get_cell_value(row, col_mapping.get('asked_by'))
        request_type = get_cell_value(row, col_mapping.get('request_type'))
        date = get_cell_value(row, col_mapping.get('date'))

        # Extract optional ID columns if detected
        about_id = get_cell_value(row, col_mapping.get('about_id'))
        from_id = get_cell_value(row, col_mapping.get('from_id'))

        # Handle date conversion
        if isinstance(date, datetime):
            feedback_date = date
        elif isinstance(date, str):
            try:
                feedback_date = datetime.fromisoformat(date)
            except ValueError:
                feedback_date = None
        else:
            feedback_date = None

        # Create feedback entry
        wd_feedback = WorkdayFeedback(
            about=about,
            from_name=from_name,
            question=question,
            feedback=feedback_text,
            asked_by=asked_by,
            request_type=request_type,
            date=feedback_date
        )

        # TODO: Add about_id and from_id columns to WorkdayFeedback model when IDs become available
        # For now, IDs are detected but not stored

        # Parse for structured feedback
        wd_feedback.parse_structured_feedback()

        # Try to add to database
        try:
            session.add(wd_feedback)
            session.flush()  # Check for constraint violations
            result.imported += 1

            if wd_feedback.is_structured:
                result.structured_count += 1
            else:
                result.generic_count += 1

        except IntegrityError:
            session.rollback()
            result.skipped_duplicates += 1

    # Commit all changes
    session.commit()
    session.close()

    # Add warnings
    if result.skipped_empty > 0:
        result.warnings.append(
            f"Skipped {result.skipped_empty} empty/incomplete rows "
            "(possibly pending feedback requests)"
        )

    if feedback_also_given_to_used:
        result.warnings.append(
            "Some entries have 'Feedback Also Given To' values - "
            "this column is not currently supported"
        )

    return result


def get_available_date_ranges(db_path='feedback.db'):
    """Get date ranges that have feedback available.

    Returns:
        List of (year, month, count) tuples sorted by date descending
    """
    session = init_db(db_path)

    # Query distinct year-month combinations with counts
    from sqlalchemy import func, extract

    results = session.query(
        extract('year', WorkdayFeedback.date).label('year'),
        extract('month', WorkdayFeedback.date).label('month'),
        func.count(WorkdayFeedback.id).label('count')
    ).filter(
        WorkdayFeedback.date.isnot(None)
    ).group_by(
        extract('year', WorkdayFeedback.date),
        extract('month', WorkdayFeedback.date)
    ).order_by(
        extract('year', WorkdayFeedback.date).desc(),
        extract('month', WorkdayFeedback.date).desc()
    ).all()

    session.close()

    return [(int(r.year), int(r.month), r.count) for r in results]


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage: python import_workday.py <xlsx_file> [db_path]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else 'feedback.db'

    print(f"Importing from {xlsx_path}...")
    result = import_workday_xlsx(xlsx_path, db_path)

    print(f"\nImport complete:")
    print(f"  Imported: {result.imported}")
    print(f"  - Structured (with tenets): {result.structured_count}")
    print(f"  - Generic (free-text): {result.generic_count}")
    print(f"  Skipped (duplicates): {result.skipped_duplicates}")
    print(f"  Skipped (empty rows): {result.skipped_empty}")

    if result.warnings:
        print(f"\nWarnings:")
        for w in result.warnings:
            print(f"  - {w}")

    if result.errors:
        print(f"\nErrors:")
        for e in result.errors:
            print(f"  - {e}")
        sys.exit(1)
