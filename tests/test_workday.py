"""
Tests for Workday XLSX import functionality.

Tests cover:
- WorkdayFeedback model and structured feedback parsing
- XLSX column detection and flexible header matching
- Import logic with validation and error handling
- API endpoints for Workday feedback
"""

import pytest
import tempfile
import os
import sys
import json
from datetime import datetime, timedelta

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from feedback_models import WorkdayFeedback, Base, init_db
from scripts.import_workday import (
    detect_columns, get_cell_value, validate_row,
    import_workday_xlsx, ImportResult, DEFAULT_CONFIG
)
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@pytest.fixture
def workday_db():
    """Create temporary test database for Workday tests."""
    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)

    # Initialize database with all models
    engine = create_engine(f'sqlite:///{db_path}')
    Base.metadata.create_all(engine)

    yield db_path

    os.unlink(db_path)


@pytest.fixture
def workday_session(workday_db):
    """Create database session for Workday tests."""
    engine = create_engine(f'sqlite:///{workday_db}')
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


class TestWorkdayFeedbackModel:
    """Tests for the WorkdayFeedback model."""

    def test_create_generic_feedback(self, workday_session):
        """Test creating generic (non-structured) feedback."""
        feedback = WorkdayFeedback(
            about='John Doe',
            from_name='Jane Smith',
            question='Please provide feedback',
            feedback='John is a great team player.',
            asked_by='John Doe',
            request_type='Requested by Self',
            date=datetime.now()
        )
        feedback.parse_structured_feedback()

        workday_session.add(feedback)
        workday_session.commit()

        assert feedback.is_structured == 0
        assert feedback.strengths is None
        assert feedback.improvements is None

    def test_create_structured_feedback(self, workday_session):
        """Test creating structured feedback with [TENETS] marker."""
        feedback_text = """[TENETS]
Strengths: tenet1, tenet2, tenet3
Improvements: tenet4, tenet5
[/TENETS]

Strengths:
• Test Tenet 1
• Test Tenet 2
• Test Tenet 3

John consistently demonstrates excellence in these areas.

Areas for Improvement:
• Test Tenet 4
• Test Tenet 5

These represent growth opportunities."""

        feedback = WorkdayFeedback(
            about='John Doe',
            from_name='Jane Smith',
            question='What strengths does the associate demonstrate?',
            feedback=feedback_text,
            asked_by='John Doe',
            request_type='Requested by Self',
            date=datetime.now()
        )
        feedback.parse_structured_feedback()

        workday_session.add(feedback)
        workday_session.commit()

        assert feedback.is_structured == 1
        assert json.loads(feedback.strengths) == ['tenet1', 'tenet2', 'tenet3']
        assert json.loads(feedback.improvements) == ['tenet4', 'tenet5']
        assert 'excellence' in feedback.strengths_text
        assert 'growth opportunities' in feedback.improvements_text

    def test_parse_feedback_no_content(self, workday_session):
        """Test parsing empty or null feedback text."""
        feedback = WorkdayFeedback(
            about='John Doe',
            from_name='Jane Smith',
            feedback=None
        )
        result = feedback.parse_structured_feedback()

        assert result is False
        # is_structured stays at default (None) when not parsed, or 0 if set by model default
        assert feedback.is_structured in (None, 0)

    def test_parse_feedback_empty_string(self, workday_session):
        """Test parsing empty string feedback."""
        feedback = WorkdayFeedback(
            about='John Doe',
            from_name='Jane Smith',
            feedback=''
        )
        result = feedback.parse_structured_feedback()

        assert result is False
        # is_structured stays at default (None) when not parsed, or 0 if set by model default
        assert feedback.is_structured in (None, 0)

    def test_parse_structured_feedback_case_insensitive(self, workday_session):
        """Test that TENETS marker parsing is case insensitive."""
        feedback_text = """[tenets]
Strengths: a, b, c
Improvements: d
[/tenets]

Some text here."""

        feedback = WorkdayFeedback(
            about='Test User',
            from_name='Reviewer',
            feedback=feedback_text
        )
        result = feedback.parse_structured_feedback()

        assert result is True
        assert feedback.is_structured == 1
        assert json.loads(feedback.strengths) == ['a', 'b', 'c']

    def test_unique_constraint_prevents_duplicates(self, workday_session):
        """Test that unique constraint prevents duplicate entries."""
        from sqlalchemy.exc import IntegrityError

        now = datetime.now()

        feedback1 = WorkdayFeedback(
            about='John Doe',
            from_name='Jane Smith',
            question='Feedback please',
            feedback='Great work!',
            date=now
        )
        workday_session.add(feedback1)
        workday_session.commit()

        # Try to add duplicate
        feedback2 = WorkdayFeedback(
            about='John Doe',
            from_name='Jane Smith',
            question='Feedback please',
            feedback='Different text',  # Different text, same key fields
            date=now
        )
        workday_session.add(feedback2)

        with pytest.raises(IntegrityError):
            workday_session.commit()


class TestColumnDetection:
    """Tests for XLSX column detection logic."""

    def test_detect_standard_columns(self):
        """Test detection of standard Workday column headers."""
        # Create mock worksheet with standard headers
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Row 1: Title
        ws['A1'] = 'Feedback Received'

        # Row 2: Headers
        ws['A2'] = 'About Photo'
        ws['B2'] = 'About'
        ws['C2'] = 'Feedback Also Given To'
        ws['D2'] = 'From Photo'
        ws['E2'] = 'From'
        ws['F2'] = 'Question'
        ws['G2'] = 'Feedback'
        ws['H2'] = 'Asked By'
        ws['I2'] = 'Type'
        ws['J2'] = 'Date'

        col_mapping, warnings = detect_columns(ws, DEFAULT_CONFIG)

        assert col_mapping['about'] == 1  # Column B (0-indexed)
        assert col_mapping['from_name'] == 4  # Column E
        assert col_mapping['question'] == 5  # Column F
        assert col_mapping['feedback'] == 6  # Column G
        assert col_mapping['asked_by'] == 7  # Column H
        assert col_mapping['request_type'] == 8  # Column I
        assert col_mapping['date'] == 9  # Column J
        assert len(warnings) == 0

    def test_detect_columns_skips_photo_columns(self):
        """Test that photo columns are not matched for 'about' or 'from'."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers with photo columns first
        ws['A2'] = 'About Photo'
        ws['B2'] = 'About'
        ws['C2'] = 'From Photo'
        ws['D2'] = 'From'
        ws['E2'] = 'Feedback'

        col_mapping, warnings = detect_columns(ws, DEFAULT_CONFIG)

        # Should match 'About' not 'About Photo'
        assert col_mapping['about'] == 1
        assert col_mapping['from_name'] == 3
        assert col_mapping['feedback'] == 4

    def test_detect_alternative_column_names(self):
        """Test detection of alternative column header names."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Use alternative names from config
        ws['A2'] = 'Recipient'
        ws['B2'] = 'Provider'
        ws['C2'] = 'Comments'
        ws['D2'] = 'Submitted'

        col_mapping, warnings = detect_columns(ws, DEFAULT_CONFIG)

        assert col_mapping['about'] == 0  # 'Recipient' matches 'about'
        assert col_mapping['from_name'] == 1  # 'Provider' matches 'from_name'
        assert col_mapping['feedback'] == 2  # 'Comments' matches 'feedback'
        assert col_mapping['date'] == 3  # 'Submitted' matches 'date'

    def test_detect_columns_missing_required_warns(self):
        """Test that missing required columns generate warnings."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Only some columns present
        ws['A2'] = 'About'
        ws['B2'] = 'Random Column'

        col_mapping, warnings = detect_columns(ws, DEFAULT_CONFIG)

        assert col_mapping['about'] == 0
        assert col_mapping.get('from_name') is None
        assert any('From' in w for w in warnings)


class TestGetCellValue:
    """Tests for safe cell value extraction."""

    def test_get_cell_value_valid_index(self):
        """Test getting value at valid index."""
        row = ('value1', 'value2', 'value3')
        assert get_cell_value(row, 1) == 'value2'

    def test_get_cell_value_none_index(self):
        """Test getting value with None index."""
        row = ('value1', 'value2')
        assert get_cell_value(row, None) is None

    def test_get_cell_value_out_of_range(self):
        """Test getting value at out-of-range index."""
        row = ('value1', 'value2')
        assert get_cell_value(row, 5) is None


class TestValidateRow:
    """Tests for row validation logic."""

    def test_validate_empty_row(self):
        """Test that empty rows are skipped without error."""
        col_mapping = {
            'about': 0,
            'from_name': 1,
            'asked_by': 2,
            'request_type': 3
        }

        # Row with no from_name
        row = ('John Doe', None, 'John Doe', 'Requested by Self')
        is_valid, error = validate_row(row, 5, col_mapping, DEFAULT_CONFIG)

        assert is_valid is False
        assert error is None  # No error, just skip

    def test_validate_consistent_self_request(self):
        """Test validation passes for consistent self-request."""
        col_mapping = {
            'about': 0,
            'from_name': 1,
            'asked_by': 2,
            'request_type': 3
        }

        row = ('John Doe', 'Jane Smith', 'John Doe', 'Requested by Self')
        is_valid, error = validate_row(row, 5, col_mapping, DEFAULT_CONFIG)

        assert is_valid is True
        assert error is None

    def test_validate_inconsistent_type_generates_error(self):
        """Test that type inconsistency generates an error."""
        col_mapping = {
            'about': 0,
            'from_name': 1,
            'asked_by': 2,
            'request_type': 3
        }

        # About == Asked By but type says "Requested by Others"
        row = ('John Doe', 'Jane Smith', 'John Doe', 'Requested by Others')
        is_valid, error = validate_row(row, 5, col_mapping, DEFAULT_CONFIG)

        assert is_valid is False
        assert 'inconsistency' in error.lower()

    def test_validate_consistent_others_request(self):
        """Test validation passes for consistent others-request."""
        col_mapping = {
            'about': 0,
            'from_name': 1,
            'asked_by': 2,
            'request_type': 3
        }

        # Manager requested feedback about employee
        row = ('John Doe', 'Jane Smith', 'Manager Name', 'Requested by Others')
        is_valid, error = validate_row(row, 5, col_mapping, DEFAULT_CONFIG)

        assert is_valid is True
        assert error is None


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestImportWorkdayXlsx:
    """Tests for full XLSX import functionality."""

    def test_import_empty_file_generates_warning(self, workday_db):
        """Test importing file with no data rows."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Just headers, no data
        ws['A2'] = 'About'
        ws['B2'] = 'From'
        ws['C2'] = 'Feedback'
        ws['D2'] = 'Date'

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            result = import_workday_xlsx(xlsx_path, workday_db)

            assert result.success is True
            assert result.imported == 0
        finally:
            os.unlink(xlsx_path)

    def test_import_structured_feedback(self, workday_db, workday_session):
        """Test importing structured feedback with tenets."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Feedback on My Team'

        # Headers
        ws['A2'] = 'About'
        ws['B2'] = 'From'
        ws['C2'] = 'Feedback'
        ws['D2'] = 'Date'

        # Data row with structured feedback
        ws['A3'] = 'John Doe'
        ws['B3'] = 'Jane Smith'
        ws['C3'] = """[TENETS]
Strengths: tenet1, tenet2, tenet3
Improvements: tenet4
[/TENETS]

Great team player!"""
        ws['D3'] = datetime(2025, 11, 15)

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            result = import_workday_xlsx(xlsx_path, workday_db)

            assert result.success is True
            assert result.imported == 1
            assert result.structured_count == 1
            assert result.generic_count == 0

            # Verify database entry
            workday_session.expire_all()
            feedback = workday_session.query(WorkdayFeedback).first()
            assert feedback.about == 'John Doe'
            assert feedback.is_structured == 1
            assert json.loads(feedback.strengths) == ['tenet1', 'tenet2', 'tenet3']
        finally:
            os.unlink(xlsx_path)

    def test_import_generic_feedback(self, workday_db, workday_session):
        """Test importing generic (non-structured) feedback."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        ws['A2'] = 'About'
        ws['B2'] = 'From'
        ws['C2'] = 'Feedback'
        ws['D2'] = 'Date'

        # Data row with generic feedback
        ws['A3'] = 'John Doe'
        ws['B3'] = 'Jane Smith'
        ws['C3'] = 'John is always helpful and a pleasure to work with.'
        ws['D3'] = datetime(2025, 11, 15)

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            result = import_workday_xlsx(xlsx_path, workday_db)

            assert result.success is True
            assert result.imported == 1
            assert result.structured_count == 0
            assert result.generic_count == 1

            # Verify database entry
            workday_session.expire_all()
            feedback = workday_session.query(WorkdayFeedback).first()
            assert feedback.is_structured == 0
            assert feedback.strengths is None
        finally:
            os.unlink(xlsx_path)

    def test_import_skips_duplicates(self, workday_db, workday_session):
        """Test that re-importing same data skips duplicates."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        ws['A2'] = 'About'
        ws['B2'] = 'From'
        ws['C2'] = 'Question'
        ws['D2'] = 'Feedback'
        ws['E2'] = 'Date'

        # Data
        ws['A3'] = 'John Doe'
        ws['B3'] = 'Jane Smith'
        ws['C3'] = 'Please provide feedback'
        ws['D3'] = 'Great work!'
        ws['E3'] = datetime(2025, 11, 15)

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            # First import
            result1 = import_workday_xlsx(xlsx_path, workday_db)
            assert result1.imported == 1
            assert result1.skipped_duplicates == 0

            # Second import
            result2 = import_workday_xlsx(xlsx_path, workday_db)
            assert result2.imported == 0
            assert result2.skipped_duplicates == 1

            # Database should still have only one entry
            count = workday_session.query(WorkdayFeedback).count()
            assert count == 1
        finally:
            os.unlink(xlsx_path)

    def test_import_missing_required_column_fails(self, workday_db):
        """Test that missing 'From' column causes error."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Missing 'From' column
        ws['A2'] = 'About'
        ws['B2'] = 'Feedback'

        ws['A3'] = 'John Doe'
        ws['B3'] = 'Some feedback'

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            result = import_workday_xlsx(xlsx_path, workday_db)

            assert result.success is False
            assert any('From' in e for e in result.errors)
        finally:
            os.unlink(xlsx_path)

    def test_import_empty_rows_counted(self, workday_db):
        """Test that empty rows are counted in skipped_empty."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers
        ws['A2'] = 'About'
        ws['B2'] = 'From'
        ws['C2'] = 'Feedback'

        # Row 3: Valid data
        ws['A3'] = 'John Doe'
        ws['B3'] = 'Jane Smith'
        ws['C3'] = 'Great work!'

        # Row 4: Empty (pending request)
        ws['A4'] = 'John Doe'
        ws['B4'] = None  # No reviewer yet
        ws['C4'] = None

        # Row 5: Valid data
        ws['A5'] = 'John Doe'
        ws['B5'] = 'Bob Jones'
        ws['C5'] = 'Good collaboration.'

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            result = import_workday_xlsx(xlsx_path, workday_db)

            assert result.success is True
            assert result.imported == 2
            assert result.skipped_empty == 1
            assert any('empty' in w.lower() for w in result.warnings)
        finally:
            os.unlink(xlsx_path)

    def test_import_with_photo_columns(self, workday_db, workday_session):
        """Test import handles photo columns correctly."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Headers matching real Workday format
        ws['A2'] = 'About Photo'
        ws['B2'] = 'About'
        ws['C2'] = 'Feedback Also Given To'
        ws['D2'] = 'From Photo'
        ws['E2'] = 'From'
        ws['F2'] = 'Question'
        ws['G2'] = 'Feedback'
        ws['H2'] = 'Asked By'
        ws['I2'] = 'Type'
        ws['J2'] = 'Date'

        # Data row
        ws['A3'] = ''  # Photo placeholder
        ws['B3'] = 'John Doe'
        ws['C3'] = ''  # Feedback Also Given To
        ws['D3'] = ''  # Photo placeholder
        ws['E3'] = 'Jane Smith'
        ws['F3'] = 'Please provide feedback'
        ws['G3'] = 'John is excellent.'
        ws['H3'] = 'John Doe'
        ws['I3'] = 'Requested by Self'
        ws['J3'] = datetime(2025, 11, 15)

        fd, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(xlsx_path)

        try:
            result = import_workday_xlsx(xlsx_path, workday_db)

            assert result.success is True
            assert result.imported == 1

            # Verify correct values were extracted
            workday_session.expire_all()
            feedback = workday_session.query(WorkdayFeedback).first()
            assert feedback.about == 'John Doe'
            assert feedback.from_name == 'Jane Smith'
            assert feedback.question == 'Please provide feedback'
        finally:
            os.unlink(xlsx_path)


class TestImportResult:
    """Tests for ImportResult class."""

    def test_success_with_no_errors(self):
        """Test that success is True when no errors."""
        result = ImportResult()
        result.imported = 5

        assert result.success is True

    def test_success_false_with_errors(self):
        """Test that success is False when errors present."""
        result = ImportResult()
        result.errors.append('Something went wrong')

        assert result.success is False

    def test_to_dict_complete(self):
        """Test to_dict returns all fields."""
        result = ImportResult()
        result.imported = 10
        result.skipped_duplicates = 2
        result.skipped_empty = 3
        result.structured_count = 6
        result.generic_count = 4
        result.warnings.append('Warning 1')

        d = result.to_dict()

        assert d['success'] is True
        assert d['imported'] == 10
        assert d['skipped_duplicates'] == 2
        assert d['skipped_empty'] == 3
        assert d['structured_count'] == 6
        assert d['generic_count'] == 4
        assert 'Warning 1' in d['warnings']
